import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import yfinance as yf
import numpy as np
from datetime import datetime, timedelta
import json
import io
import base64
import zipfile
from typing import Dict, List, Optional
import warnings
warnings.filterwarnings('ignore')

# Try to import optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    st.warning("⚠️ openpyxl not available. Excel export will be limited.")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    st.warning("⚠️ reportlab not available. PDF export will be limited.")

class ExportManager:
    """Advanced export and reporting manager"""
    
    def __init__(self):
        self.export_formats = ['CSV', 'JSON', 'Excel'] + (['PDF'] if PDF_AVAILABLE else [])
        self.report_templates = {
            'Portfolio Summary': self.generate_portfolio_summary,
            'Market Analysis': self.generate_market_analysis,
            'Risk Assessment': self.generate_risk_assessment,
            'Technical Analysis': self.generate_technical_analysis,
            'ML Predictions': self.generate_ml_predictions,
            'Custom Report': self.generate_custom_report
        }
    
    def export_dataframe(self, df: pd.DataFrame, format: str, filename: str = None) -> bytes:
        """Export dataframe in specified format"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}"
        
        if format == 'CSV':
            return df.to_csv(index=False).encode('utf-8')
        
        elif format == 'JSON':
            return df.to_json(orient='records', indent=2).encode('utf-8')
        
        elif format == 'Excel' and EXCEL_AVAILABLE:
            return self.export_to_excel(df, filename)
        
        elif format == 'PDF' and PDF_AVAILABLE:
            return self.export_to_pdf(df, filename)
        
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    def export_to_excel(self, df: pd.DataFrame, filename: str) -> bytes:
        """Export dataframe to Excel with formatting"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Export"
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def export_to_pdf(self, df: pd.DataFrame, filename: str) -> bytes:
        """Export dataframe to PDF with formatting"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not available")
        
        # Create PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Add title
        title = Paragraph(f"Financial Data Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Convert dataframe to table data
        table_data = [df.columns.tolist()] + df.values.tolist()
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output.getvalue()
    
    def generate_portfolio_summary(self, portfolio_data: List[Dict]) -> Dict:
        """Generate comprehensive portfolio summary report"""
        if not portfolio_data:
            return {"error": "No portfolio data available"}
        
        df = pd.DataFrame(portfolio_data)
        
        # Calculate summary metrics
        total_value = df['value'].sum()
        total_cost = df['cost_basis'].sum()
        total_pnl = total_value - total_cost
        total_pnl_percent = (total_pnl / total_cost) * 100 if total_cost > 0 else 0
        
        # Best and worst performers
        best_performer = df.loc[df['pnl_percent'].idxmax()]
        worst_performer = df.loc[df['pnl_percent'].idxmin()]
        
        # Sector analysis (simplified)
        sector_performance = df.groupby('symbol').agg({
            'pnl_percent': 'mean',
            'value': 'sum'
        }).sort_values('pnl_percent', ascending=False)
        
        report = {
            'title': 'Portfolio Summary Report',
            'generated_at': datetime.now().isoformat(),
            'summary': {
                'total_value': total_value,
                'total_cost': total_cost,
                'total_pnl': total_pnl,
                'total_pnl_percent': total_pnl_percent,
                'position_count': len(df)
            },
            'performance': {
                'best_performer': {
                    'symbol': best_performer['symbol'],
                    'pnl_percent': best_performer['pnl_percent']
                },
                'worst_performer': {
                    'symbol': worst_performer['symbol'],
                    'pnl_percent': worst_performer['pnl_percent']
                }
            },
            'data': df.to_dict('records')
        }
        
        return report
    
    def generate_market_analysis(self, symbols: List[str], period: str = "1mo") -> Dict:
        """Generate market analysis report"""
        market_data = {}
        analysis_results = {}
        
        for symbol in symbols:
            try:
                ticker = yf.Ticker(symbol)
                data = ticker.history(period=period)
                if not data.empty:
                    market_data[symbol] = data
                    
                    # Calculate metrics
                    returns = data['Close'].pct_change().dropna()
                    volatility = returns.std() * np.sqrt(252) * 100
                    total_return = ((data['Close'].iloc[-1] - data['Close'].iloc[0]) / data['Close'].iloc[0]) * 100
                    
                    analysis_results[symbol] = {
                        'current_price': data['Close'].iloc[-1],
                        'total_return': total_return,
                        'volatility': volatility,
                        'volume_avg': data['Volume'].mean(),
                        'high_52w': data['High'].max(),
                        'low_52w': data['Low'].min()
                    }
            except Exception as e:
                analysis_results[symbol] = {'error': str(e)}
        
        report = {
            'title': 'Market Analysis Report',
            'generated_at': datetime.now().isoformat(),
            'period': period,
            'symbols_analyzed': len(symbols),
            'analysis': analysis_results
        }
        
        return report
    
    def generate_risk_assessment(self, symbol: str, period: str = "1y") -> Dict:
        """Generate risk assessment report"""
        try:
            ticker = yf.Ticker(symbol)
            data = ticker.history(period=period)
            
            if data.empty:
                return {"error": f"No data available for {symbol}"}
            
            returns = data['Close'].pct_change().dropna()
            
            # Calculate risk metrics
            volatility = returns.std() * np.sqrt(252) * 100
            sharpe_ratio = (returns.mean() * 252) / (returns.std() * np.sqrt(252))
            
            # Maximum drawdown
            cumulative = (1 + returns).cumprod()
            running_max = cumulative.expanding().max()
            drawdown = (cumulative - running_max) / running_max
            max_drawdown = drawdown.min() * 100
            
            # Value at Risk
            var_95 = np.percentile(returns, 5) * 100
            var_99 = np.percentile(returns, 1) * 100
            
            # Risk rating
            if volatility < 20:
                risk_rating = "Low"
            elif volatility < 40:
                risk_rating = "Medium"
            else:
                risk_rating = "High"
            
            report = {
                'title': f'Risk Assessment Report - {symbol}',
                'generated_at': datetime.now().isoformat(),
                'symbol': symbol,
                'period': period,
                'risk_metrics': {
                    'volatility_annualized': volatility,
                    'sharpe_ratio': sharpe_ratio,
                    'max_drawdown': max_drawdown,
                    'var_95': var_95,
                    'var_99': var_99,
                    'risk_rating': risk_rating
                },
                'data_points': len(returns)
            }
            
            return report
            
        except Exception as e:
            return {"error": f"Error generating risk assessment: {str(e)}"}
    
    def generate_technical_analysis(self, symbol: str, period: str = "1y") -> Dict:
        """Generate technical analysis report"""
        try:
            ticker = yf.Ticker(symbol)
            data = ticker.history(period=period)
            
            if data.empty:
                return {"error": f"No data available for {symbol}"}
            
            # Calculate technical indicators
            data['SMA_20'] = data['Close'].rolling(window=20).mean()
            data['SMA_50'] = data['Close'].rolling(window=50).mean()
            
            # RSI
            delta = data['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            data['RSI'] = 100 - (100 / (1 + rs))
            
            # MACD
            data['EMA_12'] = data['Close'].ewm(span=12).mean()
            data['EMA_26'] = data['Close'].ewm(span=26).mean()
            data['MACD'] = data['EMA_12'] - data['EMA_26']
            data['MACD_Signal'] = data['MACD'].ewm(span=9).mean()
            
            # Bollinger Bands
            data['BB_Middle'] = data['Close'].rolling(window=20).mean()
            bb_std = data['Close'].rolling(window=20).std()
            data['BB_Upper'] = data['BB_Middle'] + (bb_std * 2)
            data['BB_Lower'] = data['BB_Middle'] - (bb_std * 2)
            
            # Current values
            current_price = data['Close'].iloc[-1]
            current_rsi = data['RSI'].iloc[-1]
            current_macd = data['MACD'].iloc[-1]
            current_macd_signal = data['MACD_Signal'].iloc[-1]
            
            # Technical signals
            rsi_signal = "Overbought" if current_rsi > 70 else "Oversold" if current_rsi < 30 else "Neutral"
            macd_signal = "Bullish" if current_macd > current_macd_signal else "Bearish"
            
            # Trend analysis
            sma_20 = data['SMA_20'].iloc[-1]
            sma_50 = data['SMA_50'].iloc[-1]
            trend = "Uptrend" if sma_20 > sma_50 else "Downtrend"
            
            report = {
                'title': f'Technical Analysis Report - {symbol}',
                'generated_at': datetime.now().isoformat(),
                'symbol': symbol,
                'period': period,
                'current_price': current_price,
                'technical_indicators': {
                    'rsi': current_rsi,
                    'rsi_signal': rsi_signal,
                    'macd': current_macd,
                    'macd_signal': current_macd_signal,
                    'sma_20': sma_20,
                    'sma_50': sma_50,
                    'trend': trend
                },
                'bollinger_bands': {
                    'upper': data['BB_Upper'].iloc[-1],
                    'middle': data['BB_Middle'].iloc[-1],
                    'lower': data['BB_Lower'].iloc[-1]
                }
            }
            
            return report
            
        except Exception as e:
            return {"error": f"Error generating technical analysis: {str(e)}"}
    
    def generate_ml_predictions(self, symbol: str, period: str = "1y") -> Dict:
        """Generate ML predictions report"""
        try:
            ticker = yf.Ticker(symbol)
            data = ticker.history(period=period)
            
            if data.empty:
                return {"error": f"No data available for {symbol}"}
            
            # Simple ML prediction (placeholder)
            returns = data['Close'].pct_change().dropna()
            avg_return = returns.mean()
            volatility = returns.std()
            
            # Simple prediction for next 5 days
            current_price = data['Close'].iloc[-1]
            predictions = []
            
            for i in range(1, 6):
                # Simple random walk with drift
                predicted_return = np.random.normal(avg_return, volatility)
                predicted_price = current_price * (1 + predicted_return)
                predictions.append({
                    'day': i,
                    'predicted_price': predicted_price,
                    'confidence': max(0, 100 - (volatility * 100))
                })
                current_price = predicted_price
            
            report = {
                'title': f'ML Predictions Report - {symbol}',
                'generated_at': datetime.now().isoformat(),
                'symbol': symbol,
                'current_price': data['Close'].iloc[-1],
                'predictions': predictions,
                'model_info': {
                    'type': 'Simple Random Walk',
                    'data_points': len(returns),
                    'avg_return': avg_return,
                    'volatility': volatility
                }
            }
            
            return report
            
        except Exception as e:
            return {"error": f"Error generating ML predictions: {str(e)}"}
    
    def generate_custom_report(self, data: Dict, title: str = "Custom Report") -> Dict:
        """Generate custom report from provided data"""
        report = {
            'title': title,
            'generated_at': datetime.now().isoformat(),
            'data': data
        }
        return report
    
    def create_zip_export(self, files: Dict[str, bytes], filename: str = None) -> bytes:
        """Create ZIP file with multiple exports"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"financial_export_{timestamp}.zip"
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_name, file_content in files.items():
                zip_file.writestr(file_name, file_content)
        
        zip_buffer.seek(0)
        return zip_buffer.getvalue()

def create_export_interface():
    """Create export and reporting interface"""
    st.header("📤 Export & Reporting System")
    
    # Initialize export manager
    if 'export_manager' not in st.session_state:
        st.session_state.export_manager = ExportManager()
    
    export_manager = st.session_state.export_manager
    
    # Export options
    st.subheader("📊 Data Export")
    
    # Sample data for demonstration
    sample_data = {
        'Symbol': ['AAPL', 'MSFT', 'GOOGL', 'AMZN', 'TSLA'],
        'Price': [150.25, 300.50, 2800.75, 3200.00, 250.30],
        'Change': [2.50, -5.25, 15.30, 8.75, -12.40],
        'Volume': [45000000, 25000000, 12000000, 18000000, 35000000],
        'Market Cap': [2.4e12, 2.2e12, 1.8e12, 1.6e12, 800e9]
    }
    
    df = pd.DataFrame(sample_data)
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.write("**Sample Data**")
        st.dataframe(df, use_container_width=True)
    
    with col2:
        st.write("**Export Options**")
        
        export_format = st.selectbox("Format", export_manager.export_formats)
        
        if st.button("Export Data", type="primary"):
            try:
                export_data = export_manager.export_dataframe(df, export_format)
                
                file_extension = export_format.lower()
                if export_format == 'Excel':
                    file_extension = 'xlsx'
                
                st.download_button(
                    label=f"Download {export_format}",
                    data=export_data,
                    file_name=f"financial_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}",
                    mime=f"application/{file_extension}"
                )
                
                st.success(f"✅ {export_format} export ready!")
                
            except Exception as e:
                st.error(f"❌ Export failed: {str(e)}")
    
    # Report generation
    st.subheader("📋 Report Generation")
    
    report_type = st.selectbox("Report Type", list(export_manager.report_templates.keys()))
    
    if report_type == "Portfolio Summary":
        if 'portfolio' in st.session_state and st.session_state.portfolio:
            if st.button("Generate Portfolio Report"):
                report = export_manager.generate_portfolio_summary(st.session_state.portfolio)
                
                # Display report
                st.json(report)
                
                # Export report
                report_json = json.dumps(report, indent=2)
                st.download_button(
                    label="Download Portfolio Report (JSON)",
                    data=report_json,
                    file_name=f"portfolio_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                    mime="application/json"
                )
        else:
            st.info("No portfolio data available for report generation")
    
    elif report_type == "Market Analysis":
        col1, col2 = st.columns([1, 1])
        
        with col1:
            symbols_input = st.text_input("Symbols (comma-separated)", value="AAPL,MSFT,GOOGL")
        
        with col2:
            period = st.selectbox("Period", ["1mo", "3mo", "6mo", "1y"], index=1)
        
        if st.button("Generate Market Report"):
            symbols = [s.strip().upper() for s in symbols_input.split(',')]
            report = export_manager.generate_market_analysis(symbols, period)
            
            # Display report
            st.json(report)
            
            # Export report
            report_json = json.dumps(report, indent=2)
            st.download_button(
                label="Download Market Report (JSON)",
                data=report_json,
                file_name=f"market_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
    
    elif report_type == "Risk Assessment":
        col1, col2 = st.columns([1, 1])
        
        with col1:
            symbol = st.text_input("Symbol", value="AAPL")
        
        with col2:
            period = st.selectbox("Period", ["6mo", "1y", "2y", "5y"], index=1)
        
        if st.button("Generate Risk Report"):
            report = export_manager.generate_risk_assessment(symbol, period)
            
            # Display report
            st.json(report)
            
            # Export report
            report_json = json.dumps(report, indent=2)
            st.download_button(
                label="Download Risk Report (JSON)",
                data=report_json,
                file_name=f"risk_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
    
    elif report_type == "Technical Analysis":
        col1, col2 = st.columns([1, 1])
        
        with col1:
            symbol = st.text_input("Symbol", value="AAPL")
        
        with col2:
            period = st.selectbox("Period", ["1mo", "3mo", "6mo", "1y"], index=1)
        
        if st.button("Generate Technical Report"):
            report = export_manager.generate_technical_analysis(symbol, period)
            
            # Display report
            st.json(report)
            
            # Export report
            report_json = json.dumps(report, indent=2)
            st.download_button(
                label="Download Technical Report (JSON)",
                data=report_json,
                file_name=f"technical_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
    
    elif report_type == "ML Predictions":
        col1, col2 = st.columns([1, 1])
        
        with col1:
            symbol = st.text_input("Symbol", value="AAPL")
        
        with col2:
            period = st.selectbox("Period", ["6mo", "1y", "2y"], index=1)
        
        if st.button("Generate ML Report"):
            report = export_manager.generate_ml_predictions(symbol, period)
            
            # Display report
            st.json(report)
            
            # Export report
            report_json = json.dumps(report, indent=2)
            st.download_button(
                label="Download ML Report (JSON)",
                data=report_json,
                file_name=f"ml_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
    
    # Bulk export
    st.subheader("📦 Bulk Export")
    
    if st.button("Generate All Reports", type="primary"):
        with st.spinner("Generating all reports..."):
            reports = {}
            
            # Generate all available reports
            if 'portfolio' in st.session_state and st.session_state.portfolio:
                portfolio_report = export_manager.generate_portfolio_summary(st.session_state.portfolio)
                reports['portfolio_report.json'] = json.dumps(portfolio_report, indent=2).encode('utf-8')
            
            # Market analysis for popular stocks
            market_report = export_manager.generate_market_analysis(['AAPL', 'MSFT', 'GOOGL'], '1mo')
            reports['market_report.json'] = json.dumps(market_report, indent=2).encode('utf-8')
            
            # Risk assessment for AAPL
            risk_report = export_manager.generate_risk_assessment('AAPL', '1y')
            reports['risk_report.json'] = json.dumps(risk_report, indent=2).encode('utf-8')
            
            # Technical analysis for AAPL
            technical_report = export_manager.generate_technical_analysis('AAPL', '1y')
            reports['technical_report.json'] = json.dumps(technical_report, indent=2).encode('utf-8')
            
            # ML predictions for AAPL
            ml_report = export_manager.generate_ml_predictions('AAPL', '1y')
            reports['ml_report.json'] = json.dumps(ml_report, indent=2).encode('utf-8')
            
            # Create ZIP file
            zip_data = export_manager.create_zip_export(reports)
            
            st.download_button(
                label="Download All Reports (ZIP)",
                data=zip_data,
                file_name=f"all_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime="application/zip"
            )
            
            st.success(f"✅ Generated {len(reports)} reports!")

def main():
    st.set_page_config(
        page_title="Export & Reporting System",
        page_icon="📤",
        layout="wide"
    )
    
    create_export_interface()

if __name__ == "__main__":
    main()
